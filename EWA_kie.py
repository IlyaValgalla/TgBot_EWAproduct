import telebot
from telebot import types
import pandas as pd
import openpyxl
import datetime
import threading
import time

TOKEN = '6469747259:AAFpjOqCR8ioaDpfNH94qL2x_xxoVUzPLmY'

bot = telebot.TeleBot(TOKEN)
user_data_file = 'user_data.xlsx'
running = True

user_temp_data = {}


@bot.message_handler(commands=['start'])
def start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton('Заполнить форму')
    btn2 = types.KeyboardButton('О продукции EWA')
    btn3 = types.KeyboardButton('Вся продукция')
    markup.add(btn1, btn2, btn3)
    mess = f'Приветствую тебя, {message.from_user.first_name}!\nЗаполни форму что бы ты мог следить за новинками продукции EWA'
    bot.send_message(message.chat.id, mess, reply_markup=markup)


@bot.message_handler(func=lambda message: message.text == 'Заполнить форму')
def request_name(message):
    user_temp_data[message.chat.id] = {}
    bot.send_message(message.chat.id, 'Введите ваше ФИО:')
    bot.register_next_step_handler(message, request_phone)


def request_phone(message):
    user_temp_data[message.chat.id]['ФИО'] = message.text
    bot.send_message(message.chat.id, 'Введите ваш телефон:')
    bot.register_next_step_handler(message, request_email)


def request_email(message):
    user_temp_data[message.chat.id]['Телефон'] = message.text
    bot.send_message(message.chat.id, 'Введите вашу почту:')
    bot.register_next_step_handler(message, request_nickname)


def request_nickname(message):
    user_temp_data[message.chat.id]['Почта'] = message.text
    bot.send_message(message.chat.id, 'Введите ваш никнейм:')
    bot.register_next_step_handler(message, save_user_data)


def save_user_data(message):
    user_temp_data[message.chat.id]['Никнейм'] = message.text
    save_to_excel(user_temp_data[message.chat.id])
    bot.send_message(message.chat.id, 'Спасибо! Ваши данные сохранены.')


def save_to_excel(user_data):
    try:
        workbook = openpyxl.load_workbook(user_data_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['ФИО', 'Телефон', 'Почта', 'Никнейм'])

    sheet = workbook.active
    data_row = [user_data['ФИО'], user_data['Телефон'], user_data['Почта'], user_data['Никнейм']]
    sheet.append(data_row)
    workbook.save(user_data_file)


def main_menu_markup():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton('Заполнить форму')
    btn2 = types.KeyboardButton('О продукции EWA')
    btn3 = types.KeyboardButton('Вся продукция')
    markup.add(btn1, btn2, btn3)
    return markup

@bot.message_handler(content_types=['text'])
def func(message):
    if message.chat.type == 'private':
        if(message.text == 'О продукции EWA'):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            at1 = types.KeyboardButton('OMEGA-3')
            at2 = types.KeyboardButton('BRAIN')
            at3 = types.KeyboardButton('COLLAGEN')
            at4 = types.KeyboardButton('BODYBOX')
            at5 = types.KeyboardButton('DETOX')
            at6 = types.KeyboardButton('DRAINAGE')
            at7 = types.KeyboardButton('IMMUNOPUMP')
            at8 = types.KeyboardButton('TONE')
            at9 = types.KeyboardButton('PRO SLIM')
            at10 = types.KeyboardButton('D3-Extra')
            back = types.KeyboardButton('Меню')

            markup.add(at1, at2, at3, at4, at5, at6, at7, at8, at9, at10, back)

            bot.send_message(message.chat.id, 'О продукции EWA', reply_markup=markup)

        elif(message.text == 'Вся продукция'):
            mess = f'На гугл диске ты можешь ознакомиться со всей продукцией EWA: https://drive.google.com/drive/folders/1eg5QhRPT4x_mnpb6n5J29SOValmOrl1J?usp=share_link'
            bot.send_message(message.chat.id, mess, parse_mode='html')

        elif (message.text == 'Меню'):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton('Заполнить форму')
            btn2 = types.KeyboardButton('О продукции EWA')
            btn3 = types.KeyboardButton('Вся продукция')
            markup.add(btn1, btn2, btn3)

            mess = f'Вы на главном меню'
            bot.send_message(message.chat.id, mess, parse_mode='html', reply_markup=markup)

        elif (message.text == 'OMEGA-3'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1NDRb-sYadjr4S0KP79868mG7sqdADlXW/view?usp=sharing')
        elif (message.text == 'BRAIN'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1NDRb-sYadjr4S0KP79868mG7sqdADlXW/view?usp=sharing')
        elif (message.text == 'COLLAGEN'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1qUwefC3eFJrxdi-fvtQjBnm8MW4FbzCv/view?usp=share_link')
        elif (message.text == 'BODYBOX'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1QWr8n1Ft1PNsmKQUGX1Xs1fpvlnmS88k/view?usp=share_link')
        elif (message.text == 'DETOX'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1uHLO32w_9C6BuxOB-F3b98gJ3XLj8rws/view?usp=share_link')
        elif (message.text == 'DRAINAGE'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1yBXwnU9SFmUhMgp-VPOTCBr6R2LCHttN/view?usp=share_link')
        elif (message.text == 'IMMUNOPUMP'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1ntAONR_Rl4rqDb_C5aW0IRBdAJAzOZtm/view?usp=share_link')
        elif (message.text == 'TONE'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1Eo3Hhvj6cOXIkXZXTJ9ZMappWu9zq76_/view?usp=share_link')
        elif (message.text == 'PRO SLIM'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1AM2i3HN5ggiVymalW2qsvATHVDggpIZP/view?usp=share_link')
        elif (message.text == 'D3-Extra'):
            bot.send_message(message.chat.id,
                         r'https://drive.google.com/file/d/1cVXtUwjiMah9ka84DQ3w_xdwCfFIfUv1/view?usp=share_link')


def load_user_chat_ids():
    chat_ids = []
    try:
        workbook = openpyxl.load_workbook(user_data_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            chat_id = row[0]
            chat_ids.append(chat_id)
    except FileNotFoundError:
        pass
    return chat_ids


def send_weekly_advertisement():
    chat_ids = load_user_chat_ids()

    for chat_id in chat_ids:
        try:
            bot.send_message(chat_id,
                             'Привет! Напоминаем вам о нашей продукции EWA. Посмотрите наши товары в каталоге: https://drive.google.com/drive/folders/1eg5QhRPT4x_mnpb6n5J29SOValmOrl1J?usp=share_link')
        except Exception as e:
            print(f"Error sending message to chat {chat_id}: {e}")


def schedule_weekly_advertisement():
    while running:
        current_time = datetime.datetime.now()
        weekday = current_time.weekday()
        if weekday == 4 and current_time.hour == 10:
            send_weekly_advertisement()
        time.sleep(3600)


thread = threading.Thread(target=schedule_weekly_advertisement)
thread.start()

if __name__ == '__main__':
    bot.polling(none_stop=True)
